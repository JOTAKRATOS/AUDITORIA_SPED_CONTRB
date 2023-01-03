import pandas as pd
import time
import openpyxl as xl
from openpyxl import Workbook,load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os
import streamlit as st
import zipfile


def AuditoriaSpedContribuicoes():
	
	arquivo=st.file_uploader('Selecione o arquivo do sped')
	#ArqMapaVendas=st.file_uploader('Selecione o arquivo do Mapa de Vendas')

	 # Lista NATUREZA DE OPERAÇÃO salvo na pasta do jupyter no formato tmp ou txt (pega do jeito que vem do datasul)
	def REL_NAT_OPER():

	    # criar pasta caso ela não exista
	    newpath = r'Sped Contribuições\\Arquivos' 
	    if not os.path.exists(newpath):
	        os.makedirs(newpath)

	    # abre caixa de dialogo para escolher arquivo
	    arquivo=filedialog.askopenfilename()

	    REL_NAT_OPER = pd.read_fwf(arquivo,widths=[9,31,8,2,5,5,10,7,7,9,7,6,2,12,6,5],encoding='latin-1')
	    REL_NAT_OPER=pd.DataFrame(REL_NAT_OPER)
	    REL_NAT_OPER=REL_NAT_OPER.astype(str)
	    REL_NAT_OPER=REL_NAT_OPER.rename(columns={'---------': 'COD_NAT_OP'})
	    REL_NAT_OPER=REL_NAT_OPER.rename(columns={'-------------------------------': 'DESCRICAO_NAT_OPER'})
	    REL_NAT_OPER.to_excel('Sped Contribuições\\Arquivos\\REL_NAT_OPER.xlsx',index=False)
	    #REL_NAT_OPER =pd.read_excel('C:\\Temp\\Sped Contribuições\\Arquivos\\REL_NAT_OPER.xlsx,skiprows=4')
	    messagebox.showinfo('Atualização Natureza Operação', 'Naturezas Atualizadas!')


	def ExecutaAuditoria():

		"""with st.spinner('Tabulando Mapa Vendas...'):
			# Tabula Mapa de vendas
		    MapaVendas = pd.read_fwf(ArqMapaVendas,widths=[11,40,14,14,17,12,13,15,],encoding='latin-1')
		    MapaVendas=pd.DataFrame(MapaVendas)
		    MapaVendas.columns=['COD_ITEM','Descrição_Item','MP_QtdFatur','MP_Preç_MédFatur','MP_TotalFatur','NADA','NADA1','NADA2']
		    MapaVendas.dropna(inplace=True)
		    MapaVendas=MapaVendas[['COD_ITEM','Descrição_Item','MP_QtdFatur','MP_TotalFatur','MP_Preç_MédFatur']]
		    #st.write(MapaVendas)"""

		# Mostra mensagem enquanto arquivo é processado
		with st.spinner('Aguarde enquanto o arquivo é processado...'):
	    
		    #Cria uma planilha vazia
		    #PLANILHA = pd.ExcelWriter('G://#Sistemas -ERP etc//# AUXILIUS//INCONSISTÊNCIAS EFD CONTRIBUIÇÕES.xlsx', engine='xlsxwriter')
		    arq1=[]
		    if arquivo:
		    	for line in arquivo:
		    		priv=line
		    		d = priv.decode('cp1252',errors='ignore')
		    		#d.to_string()
		    		arq1.append(d)
		    		#st.write(d)
		    	arq=pd.DataFrame(arq1)

		    df=arq[0].str.split('|',expand=True)
		    
		    # Coleta informações da empresa
		    for indice,linha in df.iterrows():
			    global CNPJ1
			    if linha[1]=='0000':
			        DT_INI=linha[6]
			        P1=DT_INI[0]+DT_INI[1]+'/'+ DT_INI[2]+DT_INI[3]+'/' + DT_INI[4]+DT_INI[5]+DT_INI[6]+DT_INI[7]
			        DT_FIN=linha[7]
			        P2=DT_FIN[0]+DT_FIN[1]+'/'+ DT_FIN[2]+DT_FIN[3]+'/' + DT_FIN[4]+DT_FIN[5]+DT_FIN[6]+DT_FIN[7]
			        NOME=linha[8]
			        CNPJ1=linha[9]
			        st.subheader('Contribuinte: ' + NOME)
			        st.subheader('Cnpj: ' + CNPJ1)
			        st.subheader('Período: ' + P1 + ' a ' + P2)
			        break

			#Cria uma planilha vazia
		    PLANILHA = pd.ExcelWriter('G://#Sistemas -ERP etc//# AUXILIUS//'+'AUD_EFD_CONTR_'+str(CNPJ1)+'-'+ str(DT_INI) + ' a ' + str(DT_FIN) +'.xlsx', engine='xlsxwriter')
		
		    #var2.set('Cnpj : ' + str(CNPJ)  + '\n Nome: ' + str(NOME) + '\n Início: ' + DT_INI + '\n Fim : ' + DT_FIN)
			    
		    # Coleta dados no REGISTRO 0200 e cria uma lista
		    global LISTA_REG_0200, TIPI, ITEM_ALIQ_ZERO
		    LISTA_REG_0200=[]
		    ALIQ_IPI_TIPI=int()
		    for indice,linha in df.iterrows():    
		        if linha[1]=='0200':
		            COD_ITEM=linha[2]
		            DESCR_ITEM=linha[3]
		            TIPO_ITEM=linha[7]
		            COD_NCM=linha[8]
		            EX_IPI=linha[9]
		            COD_GEN=linha[10]
		            COD_LST=linha[11] #Código do serviço conforme lista do Anexo I da Lei Complementar Federal nº 116/03.


		            LISTA_REG_0200.append((COD_ITEM,DESCR_ITEM,TIPO_ITEM,COD_NCM,EX_IPI,COD_LST,COD_GEN))
		    LISTA_REG_0200=pd.DataFrame(LISTA_REG_0200, columns=['COD_ITEM','DESCR_ITEM','TIPO_ITEM','COD_NCM','EX_IPI','COD_LST','COD_GEN'])
		    LISTA_REG_0200.to_excel('Sped Contribuições\\Arquivos\\LISTA_REG_0200' + '.xlsx',encoding='utf-8', index=True )
		    # Lista Itens ALIQUOTA ZERO salvo na pasta do jupyter no formato xlsx

		    ITEM_ALIQ_ZERO=pd.read_excel('Sped Contribuições\\Arquivos\\Lista_Itens_Aliq_Zero.xlsx')
		    ITEM_ALIQ_ZERO=pd.DataFrame(ITEM_ALIQ_ZERO)
		    ITEM_ALIQ_ZERO=ITEM_ALIQ_ZERO.astype(str)
		    #ITEM_ALIQ_ZERO.to_excel('c:\\temp\\' + 'ITEM_ALIQ_ZERO' + '.xlsx',encoding='utf-8', index=True )
		    
		    # Lista NCM TIPI salvo na pasta do jupyter no formato xlsx
		    TIPI=pd.read_excel('Sped Contribuições\\Arquivos\\Tipi.xlsx')
		    TIPI=pd.DataFrame(TIPI)
		    TIPI=TIPI.astype(str)
		    #TIPI.to_excel('c:\\temp\\' + 'TIPI' + '.xlsx',encoding='utf-8', index=True )
		    
		   
		    #REL_NAT_OPER 
		    REL_NAT_OPER=pd.read_excel('Sped Contribuições\\Arquivos\\REL_NAT_OPER.xlsx')
		    
		    # Lista Descrição CST IPI salvo na pasta do jupyter no formato xlsx
		    #LISTA_CST_IPI=pd.read_excel('C:\\Temp\\Sped Contribuições\\Arquivos\\Lista_Cst_IPI.xlsx')
		    #LISTA_CST_IPI=pd.DataFrame(LISTA_CST_IPI)
		    
		    # Mescla LISTA_REG_0200 com  ITEM_ALIQ_ZERO
		    #LISTA_REG_0200=pd.merge(LISTA_REG_0200,ITEM_ALIQ_ZERO, on='COD_ITEM',how='left')
		  
		    # Exclui colunas que não quero após a mescla das tabelas
		    #COLUNAS=['DESCRIÇÃO','NCM']
		   # LISTA_REG_0200.drop(COLUNAS, axis=1,inplace=True)
		    
		    # Mescla LISTA_REG_0200 com  TIPI para trazer a Aliquota IPI
		    #LISTA_REG_0200=pd.merge(LISTA_REG_0200,TIPI, on='COD_NCM',how='left')
		    #LISTA_REG_0200.to_excel('c:\\temp\\' + 'LISTA_REG_0200-tipi' + '.xlsx',encoding='utf-8', index=True )
		    
		    # MONTAGEM DE AUDITORIA

		    # Descreve na tela a etapa do processo atual....
		 
		    
		    # Define nome da planilha final e zera a lista
		    #INCONSISTÊNCIAS_EFD_CONTRIBUIÇÕES=[]
		    NOME_EXCEL='INCONS_EFD_CONTRIBUIÇÕES_BLOCO_C'
		    NOME_VERICACAO=[]


		    #----------------------------------------------------------------------------------------------------------------------------

		    # AUDITORIA - BLOCO_C

		    # Limpa algumas variáveis
		    OBSERVAÇÃO=''
		    ALIQUOTA_ZERO=''
		    CST_IPI_DESCR=str()
		    TOTAL_VL_PIS=float()
		    DIF_PIS_ITEM_NOTA=float()
		    TOTAL_VL_COFINS=float()
		    COFINS_CALCULADO=float()
		    DIF_COFINS_ITEM_NOTA=float()
		    DIF_TOTAL_PIS=float()
		    DIF_TOTAL_COFINS=float()
		    CONTADOR=0
		    SEQ=0
		    numero=0

		    DIF_PIS=0
		    DIF_COFINS=0
		    MONTAGEM_RESUMO_C=[]


		    # cria função das informações que vão aparecer no relatório
		    def CAMPOS_REL():
		    	NOME_VERICACAO.append((INCONSISTENCIA,OBSERVAÇÃO,DIF_PIS,DIF_COFINS,IND_OPER,CNPJ,\
		                               DT_DOC,DT_E_S,NUM_DOC,SER,COD_PART,COD_NAT,VL_IPI_DOC,VL_PIS_DOC,COD_NCM,COD_ITEM,DESCR_ITEM,\
		                               VL_ITEM,CST_PIS,VL_BC_PIS,ALIQ_PIS,VL_PIS,CST_COFINS,VL_BC_COFINS,\
		                               ALIQ_COFINS,VL_COFINS,VL_ICMS,VL_IPI,COD_CTA,ALIQUOTA_ZERO,CHV_NFE))

		    def CAMPOS_REL_RESUMO():
		        MONTAGEM_RESUMO_C.append((IND_OPER,CNPJ,DT_DOC,DT_E_S,NUM_DOC,SER,COD_PART,COD_NAT,VL_DOC,\
		    	VL_IPI_DOC,VL_PIS_DOC,COD_NCM,COD_ITEM,DESCR_ITEM,QTD,\
		    	VL_ITEM,CST_PIS,VL_BC_PIS,ALIQ_PIS,VL_PIS,CST_COFINS,VL_BC_COFINS,ALIQ_COFINS,VL_COFINS,\
		    	VL_ICMS,VL_IPI,COD_CTA,ALIQUOTA_ZERO,CHV_NFE))

       

		    # Inicia análise do arquivo
		    for indice,linha in df.iterrows():
		        # Coleta CNPJ
		        if linha[1]=='C010':
		            CNPJ=linha[2]

		        # Verifica diferença entre PIS dos itens com PIS da nota
		        if linha[1]!="C170" and CONTADOR ==1 and SEQ==1 :
		            DIF_PIS_ITEM_NOTA=round(DIF_PIS_ITEM_NOTA,2)
		            if DIF_PIS_ITEM_NOTA!=0:
		                INCONSISTENCIA='Diferença entre Total_Pis_Itens x Total_Pis_Nota'
		                OBSERVAÇÃO=('Total Pis Itens ' + str(TOTAL_VL_PIS) + ' Total Pis Nota ' + str(VL_PIS_DOC) + ' Dif. ' + str(DIF_PIS_ITEM_NOTA))
		                CAMPOS_REL()
		            TOTAL_VL_PIS=0
		            DIF_PIS_ITEM_NOTA=0
		            OBSERVAÇÃO=''
		            INCONSISTENCIA=''

		        # Documento de pessoa física com crédito
		        if linha[1]=='C170':
			        if len(CNPJ)==11:
			                INCONSISTENCIA='Documento de pessoa física, Verifique se é devido'
			                OBSERVAÇÃO=('Verifique se é devido')
			                CAMPOS_REL()
			                OBSERVAÇÃO=''
			                INCONSISTENCIA=''
		            
		        # Verifica diferença entre COFINS dos itens com COFINS da nota
		        if linha[1]!="C170" and CONTADOR ==1 and SEQ==1 :
		            DIF_COFINS_ITEM_NOTA=round(DIF_COFINS_ITEM_NOTA,2)
		            if DIF_COFINS_ITEM_NOTA!=0:
		                INCONSISTENCIA='Diferença entre Total_Cofins_Itens x Total_Cofins_Nota'
		                OBSERVAÇÃO=('Total Cofins Itens ' + str(TOTAL_VL_COFINS) + ' Total Cofins Nota ' + str(VL_COFINS_DOC) + ' Dif. ' + str(DIF_COFINS_ITEM_NOTA))
		                CAMPOS_REL()
		            TOTAL_VL_COFINS=0
		            DIF_COFINS_ITEM_NOTA=0
		            OBSERVAÇÃO=''
		            INCONSISTENCIA=''

		        # Coleta informações C100
		        if linha[1]=='C100':
		            CONTADOR=0
		            NUM_DOC=linha[8]
		            COD_PART=linha[4]
		            SER=linha[7]
		            IND_OPER=linha[2]
		            CHV_NFE=linha[9]
		            if IND_OPER=='0':
		                IND_OPER='Entrada'
		            else:
		                IND_OPER='Saída'
		            VL_DOC=linha[12]
		            VL_MERC_DOC=linha[16]
		            VL_IPI_DOC=linha[25]
		            VL_PIS_DOC=linha[26]
		            if VL_PIS_DOC=='':
		                VL_PIS_DOC=0
		            else:
		                VL_PIS_DOC=float(VL_PIS_DOC.replace(",", "."))
		            VL_COFINS_DOC=linha[27]
		            if VL_COFINS_DOC=='':
		                VL_COFINS_DOC=0
		            else:
		                VL_COFINS_DOC=float(VL_COFINS_DOC.replace(",", "."))
		            COD_SIT=linha[6] # Código da situação do documento fiscal, conforme a Tabela 4.1.2
		            if COD_SIT=='00':
		                COD_SIT='Documento regular'
		            elif COD_SIT=='01':
		                COD_SIT='Escrituração extemporânea de documento regular'
		            elif COD_SIT=='02':
		                COD_SIT='Documento cancelado'
		            elif COD_SIT=='03':
		                COD_SIT='Escrituração extemporânea de documento cancelado'
		            elif COD_SIT=='04':
		                COD_SIT='NF-e ou CT-e - denegado'
		            elif COD_SIT=='05':
		                COD_SIT='NF-e ou CT-e - Numeração inutilizada'
		            elif COD_SIT=='06':
		                COD_SIT='Documento Fiscal Complementar'
		            elif COD_SIT=='07':
		                COD_SIT='Escrituração extemporânea de documento complementar'
		            elif COD_SIT=='08':
		                COD_SIT='Documento Fiscal emitido com base em Regime Especial ou Norma Específica'
		            
		            DT_DOC=linha[10]
		            DT_DOC=DT_DOC[:2]+'/'+ DT_DOC[2:4]+'/'+DT_DOC[4:8]
		            DT_E_S=linha[11]
		            DT_E_S=DT_E_S[:2]+'/'+ DT_E_S[2:4]+'/'+DT_E_S[4:8]

		        # Coleta informações C170 com base na posição definida no leiaute do sped
		        if linha[1]=="C170":
		            CONTADOR=1
		            CST_ICMS=linha[10]
		            VL_ICMS=linha[15]
		            COD_NAT=linha[12]
		            COD_ITEM=str(linha[3])
		            QTD=linha[5]
		            VL_ITEM=linha[7]
		            VL_ITEM=float(VL_ITEM.replace(",", "."))
		            CST_PIS=linha[25]

		            VL_BC_PIS=(linha[26])
		            VL_BC_PIS=VL_BC_PIS.replace(",", ".")

		            ALIQ_PIS=linha[27]
		            #st.write(NUM_DOC)
		            
		            ALIQ_PIS=float(ALIQ_PIS.replace(",", "."))

		            VL_PIS=linha[30]
		            try:
		            	VL_PIS=float(VL_PIS.replace(",", "."))
		            except:
		            	VL_PIS=float(0)

		            CST_COFINS=linha[31]
		            VL_BC_COFINS=linha[32]
		            ALIQ_COFINS=linha[33]
		            ALIQ_COFINS=float(ALIQ_COFINS.replace(",", "."))
		            VL_COFINS=linha[36]
		            try:
		            	VL_COFINS=float(VL_COFINS.replace(",", "."))
		            except:
		            	VL_COFINS=float()
		            COD_CTA=linha[37]
		            VL_ICMS=linha[15]
		            VL_ICMS=float(VL_ICMS.replace(",", "."))
		            CFOP=linha[11]
		            CST_IPI=(linha[20])
		            ALIQ_IPI_DOC=linha[23]
		            try:
		             ALIQ_IPI_DOC=float(ALIQ_IPI_DOC.replace(",", "."))
		            except:
		            	ALIQ_IPI_DOC=float(0)
		            VL_IPI=linha[24]
		            VL_IPI=float(VL_IPI.replace(",", "."))
		            
		            
		            # Coleta dados na LISTA_REG_0200
		            COLETA_DADOS_0200=LISTA_REG_0200.loc[(LISTA_REG_0200['COD_ITEM']==COD_ITEM)]
		            for indice,linha in COLETA_DADOS_0200.iterrows():
		                COD_NCM=linha[3]
		                TIPO_ITEM=linha[2]
		                #ALIQ_IPI_TIPI=str(linha[5])
		                DESCR_ITEM=linha[1]
		                COD_LST=[5]
		                break
		            COLETA_DADOS_0200=ITEM_ALIQ_ZERO.loc[(ITEM_ALIQ_ZERO['COD_ITEM']==COD_ITEM)]
		            for indice,linha in COLETA_DADOS_0200.iterrows():
		                ALIQUOTA_ZERO=linha[5]
		                break
		            #COLETA_DADOS_NAT=REL_NAT_OPER.loc[(REL_NAT_OPER['Nat Oper']==COD_NAT)]
		            #for linha in REL_NAT_OPER:
		                #if linha==COD_ITEM:
		                    
		                    
		                    #DESCRICAO_NAT_OPER=linha
		                #print(REL_NAT_OPER.loc[(REL_NAT_OPER[1])
		                #break
		        
		            # Verifica DESCRIÇÃO CST do IPI lista LISTA_CST_IPI
		            #CST_IPI_DESCR=LISTA_CST_IPI.loc[(LISTA_CST_IPI['Codigo']==CST_IPI)]
		            #for indice,linha in CST_IPI_DESCR.iterrows():
		                #CST_IPI_DESCR=linha[1]


		    # REGRAS DE AUDITORIA - REGISTRO C170

		            # Item Alíquota zero com crédito/débito (não analisa crédito presumido)
		            if VL_PIS>0 and ALIQUOTA_ZERO=='SIM' and CST_PIS!='66'and COD_NAT!='2124' and COD_NAT!='5124B' and COD_NAT!='5124A' \
		            and COD_NAT!='5125' and COD_NAT!='5201F':
		                INCONSISTENCIA='Item Alíquota zero com crédito/débito'
		         
		                # Inclusão de inconsistência no relatório
		                CAMPOS_REL()
		            # limpa variáveis
		            INCONSISTENCIA=''

		            if TIPO_ITEM=='00':
		            	DESCR_TIPO_ITEM='Mercadoria para Revenda'
		            if TIPO_ITEM=='01':
		            	DESCR_TIPO_ITEM='Matéria-Prima'
		            if TIPO_ITEM=='02':
		            	DESCR_TIPO_ITEM='Embalagem'
		            if TIPO_ITEM=='03':
		            	DESCR_TIPO_ITEM='Produto em Processo'
		            if TIPO_ITEM=='04':
		            	DESCR_TIPO_ITEM='Produto Acabado'
		            if TIPO_ITEM=='05':
		            	DESCR_TIPO_ITEM='Subproduto'
		            if TIPO_ITEM=='06':
		            	DESCR_TIPO_ITEM='Produto Intermediário'
		            if TIPO_ITEM=='07':
		            	DESCR_TIPO_ITEM='Material de Uso e Consumo'
		            if TIPO_ITEM=='08':
		            	DESCR_TIPO_ITEM='Ativo Imobilizado'
		            if TIPO_ITEM=='09':
		            	DESCR_TIPO_ITEM='Serviços'
		            if TIPO_ITEM=='10':
		            	DESCR_TIPO_ITEM='Outros insumos'
		            if TIPO_ITEM=='99':
		            	DESCR_TIPO_ITEM='Outras'

		            # Tipo do Item incompatível com CFOP informado no registro C170 (Material para Revenda).
		            if CFOP=='1102' or CFOP=='2102':
		            	if TIPO_ITEM!='00':
			            	INCONSISTENCIA='Tipo do Item incompatível com CFOP (Revenda)'
			            	OBSERVAÇÃO='Tipo de Item = ' + TIPO_ITEM + ' - ' + DESCR_TIPO_ITEM
			            	CAMPOS_REL()
			            	INCONSISTENCIA=''
			            	OBSERVAÇÃO=''

			        # Tipo do Item incompatível com CFOP informado no registro C170 (Aquisição de Bens Utilizados como Insumo).
		            if CFOP=='1101' or CFOP=='2101' or CFOP=='1401' or CFOP=='2401':
		            	if TIPO_ITEM!='01' and TIPO_ITEM!='02' and TIPO_ITEM!='10':
			            	INCONSISTENCIA='Tipo do Item incompatível com CFOP (Aquisição de Bens Utilizados como Insumo)'
			            	OBSERVAÇÃO='Tipo de Item = ' + TIPO_ITEM + ' - ' + DESCR_TIPO_ITEM
			            	CAMPOS_REL()
			            	INCONSISTENCIA=''
			            	OBSERVAÇÃO=''
			        # 'Tipo do item preenchido como "Outras"
		            #if CFOP!='1101' and CFOP!='2101' and CFOP!='1401' and CFOP!='2401':
			            #if TIPO_ITEM=='99':
			            	#INCONSISTENCIA='Tipo do item preenchido como "Outras")'
			            	#OBSERVAÇÃO='Verificar se enquadramento está correto. Tipo de Item = ' + TIPO_ITEM + ' - ' + DESCR_TIPO_ITEM
			            	#CAMPOS_REL()
			            	#INCONSISTENCIA=''
			            	#OBSERVAÇÃO=''

		            DESCR_TIPO_ITEM=''
		            TIPO_ITEM=''
			        
		             # Divergência entre Base de cálculo x Cst x Tributo
		            if VL_PIS==0 and CST_PIS>'49'and CST_PIS<'68' or \
		            VL_PIS==0 and ALIQ_PIS>0 or \
		            VL_PIS>0 and CST_PIS>'66' or \
		            VL_PIS==0 and CST_PIS>'00' and CST_PIS<'04' or \
		            VL_PIS==0 and VL_BC_PIS!= "0" and VL_BC_PIS!= "0.00" and VL_BC_PIS!= "" or \
		            VL_PIS>0 and CST_PIS>'03' and CST_PIS<'14':
		                INCONSISTENCIA='Divergência entre Base de cálculo x Cst x Tributo'
		                if ALIQUOTA_ZERO=='SIM':
		                	OBSERVAÇÃO='Item Aliquota Zero'
		                else:
		                	OBSERVAÇÃO=''
		         
		                #Inclusão de inconsistência no relatório
		                CAMPOS_REL()
		            # limpa variáveis
		            INCONSISTENCIA=''
		            OBSERVAÇÃO=''

		             # Notas de Imobilizado a excluir se estiver duplicado com F130
		            if CFOP>='1551' and CFOP<='1555' or \
		            CFOP=='3551' or \
		            CFOP>='2551' and CFOP<='2555' or \
		            CFOP=='1406' or CFOP=='2406':
		                INCONSISTENCIA='Exluir Notas de Imobilizado se conter somente itens de imobilizado'
		                #Inclusão de inconsistência no relatório
		                CAMPOS_REL()
		            # limpa variáveis
		            INCONSISTENCIA=''
		            
		             # Débito Indevido
		            if VL_PIS>0 or VL_COFINS>0:
		            	if IND_OPER=='Saída' and COD_NAT!='5124B' and COD_NAT!='5124A' and COD_NAT!='5201f' and COD_NAT!='5201F':
		            		if ALIQUOTA_ZERO=='SIM' or COD_NAT=='6109':
				                INCONSISTENCIA='Débito Indevido'
				                # Inclusão de inconsistência no relatório
				                CAMPOS_REL()
		            # limpa variáveis
		            INCONSISTENCIA=''

		             # Ausência de Débito
		            if IND_OPER=='Saída' and VL_PIS==0 and ALIQUOTA_ZERO!='SIM' and COD_NCM!='11010010' and COD_ITEM!='7515' and \
		            COD_ITEM!='1020' and CFOP!='7101' and CST_PIS!='70' and COD_NAT!='5101V'and COD_NAT!='5101U'\
		            and COD_NAT!='6109' and COD_NAT=='5124B' and COD_NAT=='5124A':
		                INCONSISTENCIA='Ausência de Débito'
		                OBSERVAÇÃO=''
		                
		                # Inclusão de inconsistência no relatório
		                CAMPOS_REL()
		             # limpa variáveis
		            INCONSISTENCIA=''
		            OBSERVAÇÃO=''
		            
		             # Ausência de Débito industrIALIZAÇÃO 
		            if VL_PIS==0 and  CFOP=='5125' or VL_PIS==0 and CFOP=='5124':
		                INCONSISTENCIA='Ausência de Débito'
		                OBSERVAÇÃO='Industrialização sem Débito'
		                # Descreve na tela a etapa do processo atual....
		                time.sleep(0.1)
		                #var.set('Auditando: ' + str (INCONSISTENCIA))
		                #janela.update_idletasks()
		                # Inclusão de inconsistência no relatório
		                CAMPOS_REL()
		             # limpa variáveis
		            INCONSISTENCIA=''
		            OBSERVAÇÃO=''

		             # Ausência de Crédito
		                #CST_PIS>='50' and CST_PIS<='67'
		            if IND_OPER=='Entrada' and VL_PIS==0 and ALIQUOTA_ZERO!='SIM' and COD_ITEM!='1020'and CST_PIS!='70' and CST_PIS!='98'\
		             and CST_PIS!='73' and COD_CTA[0]!='9':
		                INCONSISTENCIA='Ausência de Crédito'
		                OBSERVAÇÃO='Analisar se crédito é devido'
		                # Inclusão de inconsistência no relatório
		                CAMPOS_REL()
		                # limpa variáveis
		            INCONSISTENCIA=''
		            OBSERVAÇÃO=''

		             # Crédito Indevido
		            if VL_PIS>0 and CST_PIS>='70' and CST_PIS<='98' and CFOP!='3551' and COD_ITEM!='7403':
		                INCONSISTENCIA='Crédito Indevido'
		                OBSERVAÇÃO='CST_PIS ' + str(CST_PIS) + ' Não gera direito à crédito'
		            	              
		                # Inclusão de inconsistência no relatório
		                CAMPOS_REL()
		            # limpa variáveis
		            INCONSISTENCIA=''
		            OBSERVAÇÃO=''

		             # Aliquota de IPI na TIPI diverge da CST informado na nota fiscal
		            #if IND_OPER=='Entrada' and ALIQ_IPI_TIPI>'0' and CST_IPI !=0 and CST_IPI !=5 and ALIQ_IPI_TIPI!='NT'and ALIQ_IPI_TIPI!='nan':
		            #    INCONSISTENCIA='Aliquota de IPI na TIPI diverge da CST informado na nota fiscal'
		            #    OBSERVAÇÃO='CST DOC = ' + ' ' + str(CST_IPI) + ' ' + str(CST_IPI_DESCR) + ' ' + 'ALIQ_IPI_DOC = ' + str(ALIQ_IPI_DOC) + ' Aliquota TIPI = ' + str(ALIQ_IPI_TIPI) 
		            #    # Descreve na tela a etapa do processo atual....
		            ##    time.sleep(0.1)
		             ##   var.set('Auditando: ' + str (INCONSISTENCIA))
		              #  janela.update_idletasks()
		                # Inclusão de inconsistência no relatório
		              #  CAMPOS_REL()
		            # limpa variáveis
		            #INCONSISTENCIA=''
		            #OBSERVAÇÃO=''

		            # Documento Regular com valor zerado
		            if COD_SIT=='00' and VL_ITEM==0 :
		                INCONSISTENCIA='Documento Regular com valor zerado'
		                # Inclusão de inconsistência no relatório
		                CAMPOS_REL()
		            # limpa variáveis
		            INCONSISTENCIA=''
		            OBSERVAÇÃO=''

		            # Aliquota Pis diferente do Padrão
		            if ALIQ_PIS!=1.65 and ALIQ_PIS!=0.5775 and ALIQ_PIS!=0 or ALIQ_COFINS!=7.6 and ALIQ_COFINS!=2.66 and ALIQ_COFINS!=0:
		            #ALIQ_COFINS!="7.6" and ALIQ_COFINS!="2.66" and ALIQ_COFINS!="7.6":
		                INCONSISTENCIA='Aliquota incorreta'
		                OBSERVAÇÃO='Analisar'
		                # Inclusão de inconsistência no relatório
		                CAMPOS_REL()
		            # limpa variáveis
		            INCONSISTENCIA=''
		            OBSERVAÇÃO=''

		           	# Aliquota Pis diferente do Padrão
		            #if ALIQ_COFINS!=7.6 and ALIQ_COFINS!=2.66 and ALIQ_COFINS!=0:
		            #    INCONSISTENCIA='Aliquota incorreta'
		            #    OBSERVAÇÃO='Analisar'
		            #    # Inclusão de inconsistência no relatório
		            #    CAMPOS_REL()
		            # limpa variáveis
		            #INCONSISTENCIA=''
		            #OBSERVAÇÃO=''
		            
		            # Crédito indevido no Administrativo ou Comercial
		            COD_CTA.split()
		            try:
		            	PRIMEIRO_NUMERO=COD_CTA[0]
		            except:
		            	PRIMEIRO_NUMERO='0'
		            if PRIMEIRO_NUMERO=='5' and VL_PIS >0 or PRIMEIRO_NUMERO=='8' and VL_PIS >0 or PRIMEIRO_NUMERO=='5' and VL_COFINS >0 or PRIMEIRO_NUMERO=='8' and VL_COFINS >0 :
		                INCONSISTENCIA='Crédito indevido no Administrativo ou Comercial'
		                OBSERVAÇÃO=''
		                # Inclusão de inconsistência no relatório
		                CAMPOS_REL()
		            # limpa variáveis
		            INCONSISTENCIA=''
		            OBSERVAÇÃO=''

		            

		            # Recálcula Pis e Cofins dos itens se a aliquota for maior que zero
		            if ALIQ_PIS>0 or ALIQ_COFINS>0:

		                if IND_OPER=='Entrada': # Tirar o ICMS da base nas devoluções de venda que tiveram pis/cofins/icms
		                    if CFOP=='1201' or CFOP=='2201' or CFOP=='1410' or CFOP=='2410' \
		                    or CFOP=='1411' or CFOP=='2411':
		                        PIS_CALCULADO=round((VL_ITEM+VL_IPI-VL_ICMS)*ALIQ_PIS/100,2)
		                        COFINS_CALCULADO=round((VL_ITEM+VL_IPI-VL_ICMS)*ALIQ_COFINS/100,2)
		                    else:
		                        PIS_CALCULADO=round((VL_ITEM+VL_IPI)*ALIQ_PIS/100,2)
		                        COFINS_CALCULADO=round((VL_ITEM+VL_IPI)*ALIQ_COFINS/100,2)

		                if IND_OPER=='Saída':
		                    PIS_CALCULADO=round((VL_ITEM-VL_ICMS)*ALIQ_PIS/100,2)
		                    COFINS_CALCULADO=round((VL_ITEM-VL_ICMS)*ALIQ_COFINS/100,2)

		               	DIF_PIS=round(PIS_CALCULADO-VL_PIS,2)
		                DIF_COFINS=round(COFINS_CALCULADO-VL_COFINS,2)
		                if DIF_PIS>0.02 or DIF_PIS<-0.02 or DIF_COFINS>0.02 or DIF_COFINS<-0.02: #Tolerância de 2 centavos na diferença
		                    DIF_BASE_PIS=round(DIF_PIS/ALIQ_PIS*100,2)
		                    DIF_BASE_COFINS=round(DIF_COFINS/ALIQ_COFINS*100,2)

		                    INCONSISTENCIA='Cálculo Incorreto Pis e/ou Cofins '
		                    OBSERVAÇÃO='CST_ICMS = ' + str(CST_ICMS) + ' VL_ICMS = ' + str(VL_ICMS) + ' Pis Sped ' + str(VL_PIS) + \
		                    ' Pis ReCalculado ' + str(PIS_CALCULADO) + \
		                    ' Dif. ' + str(DIF_PIS) + ' Dif_Base_Pis ' + str(DIF_BASE_PIS) + ' | ' + ' Cofins Sped ' \
		                    + str(VL_COFINS) + ' Cofins ReCalculado ' + str(COFINS_CALCULADO) + ' Dif. ' + str(DIF_COFINS) + \
		                    ' Dif_Base_Cofins ' + str(DIF_BASE_COFINS)
		                    #Inclusão de inconsistência no relatório
		                    CAMPOS_REL()
		                    # Descreve na tela a etapa do processo atual....
		                    #time.sleep(0.01)
		                    #var.set('Auditando: ' + str (INCONSISTENCIA))
		                    #janela.update_idletasks()
		                PIS_CALCULADO=''
		                DIF_PIS=''
		                DIF_BASE_PIS=''
		                COFINS_CALCULADO=''
		                DIF_COFINS=''
		                DIF_BASE_COFINS=''
		            
		                            
		            # MOnragem Resumo
		            #CAMPOS_REL_RESUMO()
		                
		            # limpa variáveis
		            INCONSISTENCIA=''
		            OBSERVAÇÃO=''

		            # Totaliza Pis e Cofins dos itens para compara com pis das notas
		            TOTAL_VL_PIS=round(TOTAL_VL_PIS+VL_PIS,2)
		            DIF_PIS_ITEM_NOTA=VL_PIS_DOC-TOTAL_VL_PIS
		            TOTAL_VL_COFINS=round(TOTAL_VL_COFINS+VL_COFINS,2)
		            DIF_COFINS_ITEM_NOTA=VL_COFINS_DOC-TOTAL_VL_COFINS
		            # o final dessa verificação acima está no inicio do loop
		            # em # Verifica diferença entre PIS dos itens com PIS da nota

		            # Monta resumo
		            CAMPOS_REL_RESUMO()        
		            
		        # limpa variáveis
		        ALIQUOTA_ZERO=''
		        INCONSISTENCIA=''
		        # SEQ=1 serve para calcular a diferença entre PIS dos itens com PIS da nota
		        SEQ=1
		        
		        
		        
		        # Limita o loop até o último registro do Bloco_C
		        if linha[1]=="C990":
		            break 
		 	    
		    # MONTA RELATÓRIO FINAL - BLOCO_C
		    NOME_VERICACAO=pd.DataFrame(NOME_VERICACAO,columns=['INCONSISTÊNCIA','OBSERVAÇÃO','DIF_PIS','DIF_COFINS','IND_OPER',\
		                                                        'CNPJ','DT_DOC','DT_E_S','NUM_DOC','SER','COD_PART','COD_NAT',\
		                                                        'VL_IPI_DOC',\
		                                                        'VL_PIS_DOC','COD_NCM','COD_ITEM','DESCR_ITEM','VL_ITEM','CST_PIS',\
		                                                        'VL_BC_PIS','ALIQ_PIS','VL_PIS','CST_COFINS','VL_BC_COFINS',\
		                                                        'ALIQ_COFINS','VL_COFINS','VL_ICMS','VL_IPI','COD_CTA',\
		                                                        'ALIQUOTA_ZERO','CHV_NFE'])
		    MONTAGEM_RESUMO_C=pd.DataFrame(MONTAGEM_RESUMO_C,columns=['IND_OPER','CNPJ','DT_DOC','DT_E_S','NUM_DOC','SER','COD_PART','COD_NAT',\
		    	'VL_DOC','VL_IPI_DOC','VL_PIS_DOC','COD_NCM','COD_ITEM','DESCR_ITEM','QTD',\
		    	'VL_ITEM','CST_PIS','VL_BC_PIS','ALIQ_PIS','VL_PIS','CST_COFINS','VL_BC_COFINS','ALIQ_COFINS','VL_COFINS',\
		    	'VL_ICMS','VL_IPI','COD_CTA','ALIQUOTA_ZERO','CHV_NFE'])


		    # Remove linhas duplicadas
		    NOME_VERICACAO=NOME_VERICACAO.drop_duplicates()
		    NOME_VERICACAO.to_excel(PLANILHA, sheet_name='BLOCO_C',index=False)

		    NOME_VERICACAO.to_html()
		    
		    MONTAGEM_RESUMO_C=MONTAGEM_RESUMO_C.drop_duplicates()
		    

		#------------------------------------------------------------------------------------------------
		    # AUDITORIA - BLOCO_D - Documentos Fiscais – II - Serviços (ICMS)
		    
		    # Define nome da planilha final e zera a lista
		    NOME_EXCEL='INCONS_EFD_CONTRIBUIÇÕES_BLOCO_D'
		    RELATORIO_BLOCO_D=[]
		    RELATORIO_BLOCO_D_RESUMO=[]
		    # cria função das informações que vão aparecer no relatório
		    def CAMPOS_REL_BLOCO_D():
		        RELATORIO_BLOCO_D.append((INCONSISTENCIA,OBSERV,CNPJ,IND_EMIT,COD_PART,COD_SIT,SER,NUM_DOC,CHV_CTE,\
		                                  DT_DOC,DT_A_P,VL_DOC,VL_DESC,IND_FRT,VL_SERV,COD_CTA,IND_NAT_FRT,\
		                                  VL_ITEM,CST_PIS,NAT_BC_CRED,VL_BC_PIS,ALIQ_PIS,VL_PIS,VL_BC_COFINS,\
		                                  ALIQ_COFINS,VL_COFINS,COD_CTA))
		    def CAMPOS_REL_BLOCO_D_RESUMO():
		        RELATORIO_BLOCO_D_RESUMO.append((CNPJ,IND_EMIT,COD_PART,COD_SIT,SER,NUM_DOC,CHV_CTE,\
		                                  DT_DOC,DT_A_P,VL_DOC,VL_DESC,IND_FRT,VL_SERV,COD_CTA,IND_NAT_FRT,\
		                                  VL_ITEM,CST_PIS,NAT_BC_CRED,VL_BC_PIS,ALIQ_PIS,VL_PIS,VL_BC_COFINS,\
		                                  ALIQ_COFINS,VL_COFINS,COD_CTA))
		  
		    # Inicia análise do arquivo
		    for indice,linha in df.iterrows():
		        # Registro D010: Identificação do Estabelecimento
		        if linha[1]=='D010':
		            CNPJ=linha[2]
		        # Registro D100: Aquisição de Serviços de Transporte 
		        if linha[1]=='D100':
		            IND_EMIT=linha[3]
		            if IND_EMIT=='0':
		                IND_EMIT='Emissão Própria'
		            else:
		                IND_EMIT='Emissão por Terceiros'
		            COD_PART=linha[4]
		            COD_SIT=linha[6]
		            SER=linha[7]
		            NUM_DOC=linha[9]
		            CHV_CTE=linha[10]
		            DT_DOC=linha[11] # emissão
		            DT_A_P=linha[12] # entrada
		            VL_DOC=linha[15] # Valor total do documento fiscal
		            VL_DESC=linha[16]
		            IND_FRT=linha[17]
		            VL_SERV=linha[18] # Valor total da prestação de serviço
		            COD_CTA=linha[23]
		        # Registro D101: Complemento do Documento de Transporte PIS
		        if linha[1]=='D101':
		            IND_NAT_FRT=linha[2]
		            VL_ITEM=linha[3]
		            VL_ITEM=float(VL_ITEM.replace(",", "."))
		            CST_PIS=linha[4]
		            NAT_BC_CRED=linha[5]
		            VL_BC_PIS=linha[6]
		            VL_BC_PIS=float(VL_BC_PIS.replace(",", "."))
		            ALIQ_PIS=linha[7]
		            ALIQ_PIS=float(ALIQ_PIS.replace(",", "."))
		            VL_PIS=linha[8]
		            VL_PIS=float(VL_PIS.replace(",", "."))

		            CAMPOS_REL_BLOCO_D_RESUMO()
		            
		             # ReCálculo Pis
		            RECALCULO_PIS=round((VL_ITEM*1.65)/100,2)
		            DIF_RECALCULO_PIS=round(RECALCULO_PIS-VL_PIS,2)
		            if DIF_RECALCULO_PIS>0.02 and IND_NAT_FRT<'3' or DIF_RECALCULO_PIS<-0.02 and IND_NAT_FRT<'3' :
		                INCONSISTENCIA='Cálculo Pis Incorreto'
		                OBSERV=str(VL_ITEM) + ' * ' + str(ALIQ_PIS) + ' = ' + str(RECALCULO_PIS) + ' Dif: ' + str(DIF_RECALCULO_PIS)
		                VL_BC_COFINS=''
		                ALIQ_COFINS=''
		                VL_COFINS=''
		                CAMPOS_REL_BLOCO_D()
		                INCONSISTENCIA=''
		                OBSERV=''
		            RECALCULO_PIS=''
		            DIF_RECALCULO_PIS=''
		            
		            # Cst x Aliq x Pis
		            if ALIQ_PIS!=1.65 and VL_PIS>0 or CST_PIS !='56' and VL_PIS>0 or CST_PIS =='56' and VL_PIS==0 :
		                INCONSISTENCIA='Divergência entre Cst x Aliq x Pis'
		                OBSERV=''
		                VL_BC_COFINS=''
		                ALIQ_COFINS=''
		                VL_COFINS=''
		                CAMPOS_REL_BLOCO_D()
		                INCONSISTENCIA=''
		                OBSERV=''

		                
		    #Registro D100: Aquisição de Serviços de Transporte
		    for indice,linha in df.iterrows():
		        if linha[1]=='D010':
		            CNPJ=linha[2]
		        # Registro D100: Aquisição de Serviços de Transporte 
		        if linha[1]=='D100':
		            IND_EMIT=linha[3]
		            if IND_EMIT=='0':
		                IND_EMIT='Emissão Própria'
		            else:
		                IND_EMIT='Emissão por Terceiros'
		            COD_PART=linha[4]
		            COD_SIT=linha[6]
		            SER=linha[7]
		            NUM_DOC=linha[9]
		            CHV_CTE=linha[10]
		            DT_DOC=linha[11] # emissão
		            DT_A_P=linha[12] # entrada
		            VL_DOC=linha[15] # Valor total do documento fiscal
		            VL_DESC=linha[16]
		            IND_FRT=linha[17]
		            VL_SERV=linha[18] # Valor total da prestação de serviço
		            COD_CTA=linha[23]
		      
		        # Registro D105: Complemento do Documento de Transporte COFINS
		        if linha[1]=='D105':
		            IND_NAT_FRT=linha[2]
		            VL_ITEM=linha[3]
		            VL_ITEM=float(VL_ITEM.replace(",", "."))
		            CST_COFINS=linha[4]
		            NAT_BC_CRED=linha[5]
		            VL_BC_COFINS=linha[6]
		            VL_BC_COFINS=float(VL_BC_COFINS.replace(",", "."))
		            ALIQ_COFINS=linha[7]
		            ALIQ_COFINS=float(ALIQ_COFINS.replace(",", "."))
		            VL_COFINS=linha[8]
		            VL_COFINS=float(VL_COFINS.replace(",", "."))
		            COD_CTA=linha[9]
		            
		             # ReCálculo Cofins
		            RECALCULO_COFINS=round((VL_ITEM*7.6)/100,2)
		            DIF_RECALCULO_COFINS=round(RECALCULO_COFINS-VL_COFINS,2)
		            if DIF_RECALCULO_COFINS>0.02 and IND_NAT_FRT<'3' or DIF_RECALCULO_COFINS<-0.02 and IND_NAT_FRT<'3':
		                INCONSISTENCIA='Cálculo Cofins Incorreto'
		                OBSERV=str(VL_ITEM) + ' * ' + str(ALIQ_COFINS) + ' = ' + str(RECALCULO_COFINS) + ' Dif: ' + str(DIF_RECALCULO_COFINS)
		                VL_BC_PIS=''
		                ALIQ_PIS=''
		                VL_PIS=''
		                CAMPOS_REL_BLOCO_D()
		                INCONSISTENCIA=''
		                OBSERV=''
		            RECALCULO_COFINS=''
		            DIF_RECALCULO_COFINS=''
		            
		            # Cst x Aliq x Cofins
		            if ALIQ_COFINS!=7.6 and VL_COFINS>0 or CST_COFINS !='56' and VL_COFINS>0 or CST_COFINS =='56' and VL_COFINS==0 :
		                INCONSISTENCIA='Divergência entre Cst x Aliq x Cofins'
		                OBSERV= str(ALIQ_COFINS) +' / '+ str(VL_COFINS) +' / '+ str(CST_COFINS)
		                VL_BC_PIS=''
		                ALIQ_PIS=''
		                VL_PIS=''
		                CAMPOS_REL_BLOCO_D()
		                INCONSISTENCIA=''
		                OBSERV=''

		    # MONTA RELATÓRIO FINAL - BLOCO_D - Cabeçalhos
		    RELATORIO_BLOCO_D=pd.DataFrame(RELATORIO_BLOCO_D,columns=['INCONSISTÊNCIA','OBSERVACAO','CNPJ','IND_EMIT',\
		                                                              'COD_PART','COD_SIT','SER','NUM_DOC','CHV_CTE',\
		                                                              'DT_DOC','DT_A_P','VL_DOC','VL_DESC','IND_FRT',\
		                                                              'VL_SERV','COD_CTA','IND_NAT_FRT','VL_ITEM','CST_PIS',\
		                                                              'NAT_BC_CRED','VL_BC_PIS','ALIQ_PIS','VL_PIS',\
		                                                              'VL_BC_COFINS','ALIQ_COFINS','VL_COFINS','COD_CTA'])
		    RELATORIO_BLOCO_D_RESUMO=pd.DataFrame(RELATORIO_BLOCO_D_RESUMO,columns=['CNPJ','IND_EMIT',\
		                                                              'COD_PART','COD_SIT','SER','NUM_DOC','CHV_CTE',\
		                                                              'DT_DOC','DT_A_P','VL_DOC','VL_DESC','IND_FRT',\
		                                                              'VL_SERV','COD_CTA','IND_NAT_FRT','VL_ITEM','CST_PIS',\
		                                                              'NAT_BC_CRED','VL_BC_PIS','ALIQ_PIS','VL_PIS',\
		                                                              'VL_BC_COFINS','ALIQ_COFINS','VL_COFINS','COD_CTA'])
		    # Remove linhas duplicadas
		    RELATORIO_BLOCO_D=RELATORIO_BLOCO_D.drop_duplicates()
		    #RELATORIO_BLOCO_D.to_excel('c:\\temp\\' + NOME_EXCEL + '.xlsx',encoding='utf-8', index=True )
		    RELATORIO_BLOCO_D.to_excel(PLANILHA, sheet_name='BLOCO_D',index=False)
		    RELATORIO_BLOCO_D_RESUMO.to_excel(PLANILHA, sheet_name='RESUMO_D',index=False)
		    #PLANILHA.save()
		    
		    #FIM ANALISE BLOCO D
		    
		    
#--------------------------------------------------------------------------------------------------------------------------------------------      
		    # ANALISE REGISTRO C500
		
		    
		    # Define nome da planilha final e zera a lista
		    NOME_EXCEL='INCONS_EFD_CONTRIBUIÇÕES_BLOCO_C500'
		    RELATORIO_BLOCO_C500=[]
		    # cria função das informações que vão aparecer no relatório
		    def CAMPOS_REL_BLOCO_C500():
		        RELATORIO_BLOCO_C500.append((INCONSISTENCIA,OBSERV,REG,COD_PART,COD_SIT,SER,NUM_DOC))

		    for indice,linha in df.iterrows():
		        #Registro C500: Nota Fiscal/Conta de Energia Elétrica 
		        if linha[1]=="C500":
		            REG=linha[1]
		            COD_PART=linha[2]
		            COD_SIT=linha[4]
		            SER=linha[5]
		            NUM_DOC=linha[7]
		            DT_DOC=linha[8]
		            DT_ENT=linha[9]
		            VL_DOC=linha[10]
		            VL_PIS_NOTA=linha[13]
		            VL_COFINS_NOTA=linha[14]
		            
		        #Registro C501: Complemento da Operação (Códigos 06, 28 e 29) – PIS/Pasep
		        if linha[1]=="C501":
		            REG=linha[1]
		            CST_PIS=linha[2]
		            VL_ITEM=linha[3]
		            NAT_BC_CRED=linha[4]
		            VL_BC_PIS=linha[5]
		            ALIQ_PIS=linha[6]
		            VL_PIS_ITEM=linha[7]
		            COD_CTA=linha[8]
		            
		            if VL_PIS_NOTA!=VL_PIS_ITEM :
		                INCONSISTENCIA='Total Pis do Item Diferente do Total Pis da Nota'
		                OBSERV='VL_PIS_NOTA = ' + str(VL_PIS_NOTA) + ' / ' +  'VL_PIS_ITEM = ' + str(VL_PIS_ITEM)
		                CAMPOS_REL_BLOCO_C500()
		                INCONSISTENCIA=''
		                OBSERV=''
		            if VL_PIS_ITEM=='0' :
		                INCONSISTENCIA='Ausência Crédito Energia Elétrica'
		                OBSERV=''
		                CAMPOS_REL_BLOCO_C500()
		                INCONSISTENCIA=''
		                OBSERV=''
		            
		        #Registro C505: Complemento da Operação (Códigos 06, 28 e 29) – Cofins
		        if linha[1]=="C505":
		            REG=linha[1]
		            CST_COFINS=linha[2]
		            VL_COFINS_ITEM=linha[3]
		            NAT_BC_CRED_COF=linha[4]
		            VL_BC_COFINS=linha[5]
		            ALIQ_COFINS=linha[6]
		            VL_COFINS_ITEM=linha[7]
		            COD_CTA_COF=linha[8]
		        
		            #REGRAS
		            if VL_COFINS_NOTA!=VL_COFINS_ITEM:
		                INCONSISTENCIA='Total Cofins do Item Diferente do Total Cofins da Nota'
		                OBSERV='VL_COFINS_NOTA = ' + str(VL_COFINS_NOTA) + ' / ' +  'VL_COFINS_ITEM = ' + str(VL_COFINS_ITEM)
		                CAMPOS_REL_BLOCO_C500()
		                INCONSISTENCIA=''
		                OBSERV=''

		    # MONTA RELATÓRIO FINAL - BLOCO_D - Cabeçalhos
		    RELATORIO_BLOCO_C500=pd.DataFrame(RELATORIO_BLOCO_C500,columns=['INCONSISTÊNCIA','OBSERVACAO','REG','COD_PART',\
		                                                                    'COD_SIT','SER','NUM_DOC'])
		    # Remove linhas duplicadas
		    RELATORIO_BLOCO_C500=RELATORIO_BLOCO_C500.drop_duplicates()
		    #RELATORIO_BLOCO_C500.to_excel('c:\\temp\\' + NOME_EXCEL + '.xlsx',encoding='utf-8', index=True )
		    RELATORIO_BLOCO_C500.to_excel(PLANILHA, sheet_name='BLOCO_C500',index=False)
		    #PLANILHA.save()  
		    
#--------------------------------------------------------------------------------------------------------------------------------------------
      
		    # ANALISE REGISTRO F100
		
		    
		    # Define nome da planilha final e zera a lista
		    NOME_EXCEL='INCONS_EFD_CONTRIBUIÇÕES_BLOCO_F100'
		    RELATORIO_BLOCO_F100=[]
		    RELATORIO_BLOCO_F100_RESUMO=[]
		    # cria função das informações que vão aparecer no relatório
		    def CAMPOS_REL_BLOCO_F100():
		        RELATORIO_BLOCO_F100.append((INCONSISTENCIA,OBSERV,CNPJ,IND_OPER,COD_PART,DT_OPER,VL_OPER,CST_PIS,VL_BC_PIS,\
		        	ALIQ_PIS,VL_PIS,CST_COFINS,VL_BC_COFINS,ALIQ_COFINS,VL_COFINS,NAT_BC_CRED,IND_ORIG_CRED,COD_CTA,\
		        	COD_CCUS,DESC_DOC_OPER))
		    def CAMPOS_REL_BLOCO_F100_RESUMO():
		        RELATORIO_BLOCO_F100_RESUMO.append((CNPJ,IND_OPER,COD_PART,DT_OPER,VL_OPER,CST_PIS,VL_BC_PIS,\
		        	ALIQ_PIS,VL_PIS,CST_COFINS,VL_BC_COFINS,ALIQ_COFINS,VL_COFINS,NAT_BC_CRED,IND_ORIG_CRED,COD_CTA,\
		        	COD_CCUS,DESC_DOC_OPER))

		    for indice,linha in df.iterrows():
		        #Registro F100: 
		        if linha[1]=="F010":
		        	CNPJ=linha[2]

		        if linha[1]=="F100":
		        	IND_OPER=linha[2]
		        	COD_PART=linha[3]
		        	DT_OPER=linha[5]
		        	VL_OPER=linha[6]
		        	CST_PIS=linha[7]
		        	VL_BC_PIS=linha[8]
		        	ALIQ_PIS=linha[9]
		        	VL_PIS=linha[10]
		        	CST_COFINS=linha[11]
		        	VL_BC_COFINS=linha[12]
		        	ALIQ_COFINS=linha[13]
		        	VL_COFINS=linha[14]
		        	NAT_BC_CRED=linha[15]
		        	IND_ORIG_CRED=linha[16]
		        	COD_CTA=linha[17]
		        	COD_CCUS=linha[18]
		        	DESC_DOC_OPER=linha[19]

		        	CAMPOS_REL_BLOCO_F100_RESUMO()

		        	for indice,linha in df.iterrows():
		        		if linha[1]=="0150":
		        			if COD_PART==linha[2]:
		        				CPF_participante=linha[6]
		        				NOME_participante=linha[3]
					        	if len(CPF_participante)==11:
					        		INCONSISTENCIA='Documento de Pessoa Fisica'
					        		OBSERV='CPF_participante ' + CPF_participante + '-' + NOME_participante
					        		CAMPOS_REL_BLOCO_F100()
					        		INCONSISTENCIA=''
					        		OBSERV=''
					        	break

		            

		    # MONTA RELATÓRIO FINAL 
		    RELATORIO_BLOCO_F100=pd.DataFrame(RELATORIO_BLOCO_F100,columns=['INCONSISTENCIA','OBSERV','CNPJ','IND_OPER','COD_PART','DT_OPER','VL_OPER',\
		    	'CST_PIS','VL_BC_PIS','ALIQ_PIS','VL_PIS','CST_COFINS','VL_BC_COFINS','ALIQ_COFINS','VL_COFINS','NAT_BC_CRED','IND_ORIG_CRED','COD_CTA',\
		        	'COD_CCUS','DESC_DOC_OPER'])

		    RELATORIO_BLOCO_F100_RESUMO=pd.DataFrame(RELATORIO_BLOCO_F100_RESUMO,columns=['CNPJ','IND_OPER','COD_PART','DT_OPER','VL_OPER',\
		    	'CST_PIS','VL_BC_PIS','ALIQ_PIS','VL_PIS','CST_COFINS','VL_BC_COFINS','ALIQ_COFINS','VL_COFINS','NAT_BC_CRED','IND_ORIG_CRED','COD_CTA',\
		        	'COD_CCUS','DESC_DOC_OPER'])
		    # Remove linhas duplicadas
		    RELATORIO_BLOCO_F100=RELATORIO_BLOCO_F100.drop_duplicates()
		    RELATORIO_BLOCO_F100.to_excel(PLANILHA, sheet_name='BLOCO_F100',index=False)
		    RELATORIO_BLOCO_F100_RESUMO.to_excel(PLANILHA, sheet_name='RESUMO_F100',index=False)
		    # Monta resumo Bloco F100 

		    
#--------------------------------------------------------------------------------------------------------------------------------------------      
		    # ANALISE Registro F120: Bens Incorporados ao Ativo Imobilizado
		    # Operações Geradoras de Créditos com Base nos Encargos de Depreciação e Amortização
		
		    
		    # Define nome da planilha final e zera a lista
		    NOME_EXCEL='INCONS_EFD_CONTRIBUIÇÕES_BLOCO_F120'
		    RELATORIO_BLOCO_F120=[]
		    # cria função das informações que vão aparecer no relatório
		    def CAMPOS_REL_BLOCO_F120():
		        RELATORIO_BLOCO_F120.append((INCONSISTENCIA,OBSERV,REG,CNPJ,IND_ORIG_CRED,DESC_BEM_IMOB,VL_BC_PIS,\
		                                     ALIQ_PIS,VL_PIS))

		    for indice,linha in df.iterrows():
		        #Registro C500: Nota Fiscal/Conta de Energia Elétrica 
		        if linha[1]=="F010":
		            REG=linha[1]
		            CNPJ=linha[2]
		        if linha[1]=="F120":
		            REG=linha[1]
		            NAT_BC_CRED=linha[2]
		            IDENT_BEM_IMOB=linha[2]
		            IND_ORIG_CRED=linha[4]
		            IND_UTIL_BEM_IMOB=linha[5]
		            VL_OPER_DEP=linha[6]
		            PARC_OPER_NAO_BC_CRED=linha[7]
		            CST_PIS=linha[8]
		            VL_BC_PIS=linha[9]
		            VL_BC_PIS=float(VL_BC_PIS.replace(",", "."))
		            ALIQ_PIS=linha[10]
		            ALIQ_PIS=float(ALIQ_PIS.replace(",", "."))
		            VL_PIS=linha[11]
		            VL_PIS=float(VL_PIS.replace(",", "."))
		            CST_COFINS=linha[12]
		            VL_BC_COFINS=linha[13]
		            VL_BC_COFINS=float(VL_BC_COFINS.replace(",", "."))
		            ALIQ_COFINS=linha[14]
		            ALIQ_COFINS=float(ALIQ_COFINS.replace(",", "."))
		            VL_COFINS=linha[15]
		            VL_COFINS=float(VL_COFINS.replace(",", "."))
		            COD_CTA=linha[16]
		            COD_CCUS=linha[17]
		            DESC_BEM_IMOB=linha[18]
		            
		            #REGRAS
		            # RECALCULO PIS
		            RECALC_PIS_F120=round((VL_BC_PIS * ALIQ_PIS / 100),2)
		            DIF_RECALC_PIS_F120=round(RECALC_PIS_F120-VL_PIS,2)
		            if DIF_RECALC_PIS_F120>0.1 or DIF_RECALC_PIS_F120<-0.1:
		                INCONSISTENCIA='Diferença Cálculo Pis'
		                OBSERV='Valor Sped = ' + str(VL_PIS)+' / ' + 'Valor Recalculado = ' + str(RECALC_PIS_F120) +' / ' + ' Dif = ' + str(DIF_RECALC_PIS_F120)
		                CAMPOS_REL_BLOCO_F120()
		                INCONSISTENCIA=''
		                OBSERV=''
		            # RECALCULO COFINS
		            RECALC_COFINS_F120=round((VL_BC_COFINS * ALIQ_COFINS / 100),2)
		            DIF_RECALC_COFINS_F120=round(RECALC_COFINS_F120-VL_COFINS,2)
		            if DIF_RECALC_COFINS_F120>0.1 or DIF_RECALC_COFINS_F120<-0.1:
		                INCONSISTENCIA='Diferença Cálculo COfins'
		                OBSERV='Valor Sped = ' + str(VL_COFINS)+' / ' + 'Valor Recalculado = ' + str(RECALC_COFINS_F120)+' / ' + ' Dif = ' + str(DIF_RECALC_COFINS_F120)
		                CAMPOS_REL_BLOCO_F120()
		                INCONSISTENCIA=''
		                OBSERV=''
		            # ORIGEM BEM X ALIQUOTA
		            if IND_ORIG_CRED=='0' and ALIQ_PIS!=1.65 or IND_ORIG_CRED=='0' and ALIQ_COFINS!=7.6:
		                INCONSISTENCIA='Bem Nacional com Aliquota Divergente'
		                OBSERV='Verifique se a aliquota está correta para bem nacional'
		                CAMPOS_REL_BLOCO_F120()
		                INCONSISTENCIA=''
		                OBSERV=''
		             # ORIGEM BEM X ALIQUOTA
		            if IND_ORIG_CRED=='1' and ALIQ_PIS==1.65 or IND_ORIG_CRED=='1' and ALIQ_COFINS==7.6:
		                INCONSISTENCIA='Bem Importado com Aliquota Divergente'
		                OBSERV='Verifique se a aliquota está correta para bem Importado'
		                CAMPOS_REL_BLOCO_F120()
		                INCONSISTENCIA=''
		                OBSERV=''
		            # CST X ALIQUOTA
		            if CST_PIS==56 and ALIQ_PIS==0 or CST_PIS==56 and ALIQ_COFINS==0 :
		                INCONSISTENCIA='Aliquota incorreta para CST gerador de crédito'
		                OBSERV=''
		                CAMPOS_REL_BLOCO_F120()
		                INCONSISTENCIA=''
		                OBSERV=''

		            
		  # MONTA RELATÓRIO FINAL - BLOCO_D - Cabeçalhos
		    RELATORIO_BLOCO_F120=pd.DataFrame(RELATORIO_BLOCO_F120,columns=['INCONSISTÊNCIA','OBSERVACAO','REG','CNPJ',\
		                                                                    'IND_ORIG_CRED','DESC_BEM_IMOB','VL_BC_PIS',\
		                                                                    'ALIQ_PIS','VL_PIS'])
		    # Remove linhas duplicadas
		    RELATORIO_BLOCO_F120=RELATORIO_BLOCO_F120.drop_duplicates()
		    RELATORIO_BLOCO_F120.to_excel(PLANILHA, sheet_name='BLOCO_F120',index=False)
		    #RELATORIO_BLOCO_F120.to_excel(PLANILHA, sheet_name='BLOCO_F120')
		    #PLANILHA.save()
		    
		    #-------------------------------------------------------------------------------------------------      
		    # Registro F130: Bens Incorporados ao Ativo Imobilizado 
		    #– Operações Geradoras de Créditos com Base no Valor de Aquisição/Contribuição
		    
		    # Define nome da planilha final e zera a lista
		    NOME_EXCEL='INCONS_EFD_CONTRIBUIÇÕES_BLOCO_F130'
		    RELATORIO_BLOCO_F130=[]
		    # cria função das informações que vão aparecer no relatório
		    def CAMPOS_REL_BLOCO_F130():
		        RELATORIO_BLOCO_F130.append((INCONSISTENCIA,OBSERV,REG,CNPJ,IND_ORIG_CRED,DESC_BEM_IMOB,VL_BC_COFINS,ALIQ_COFINS,\
		                                     VL_COFINS,VL_BC_PIS,ALIQ_PIS,VL_PIS))

		    for indice,linha in df.iterrows():
		        #Registro C500: Nota Fiscal/Conta de Energia Elétrica 
		        if linha[1]=="F010":
		            REG=linha[1]
		            CNPJ=linha[2]
		        if linha[1]=="F130":
		            REG=linha[1]
		            NAT_BC_CRED=linha[2]
		            IDENT_BEM_IMOB=linha[3]
		            IND_ORIG_CRED=linha[4]
		            IND_UTIL_BEM_IMOB=linha[5]
		            MES_OPER_AQUIS=linha[6]
		            VL_OPER_AQUIS=linha[7]
		            PARC_OPER_NAO_BC_CRED=linha[8]
		            VL_BC_CRED=linha[9]
		            IND_NR_PARC=linha[10]
		            CST_PIS=linha[11]
		            VL_BC_PIS=linha[12]
		            VL_BC_PIS=float(VL_BC_PIS.replace(",", "."))
		            ALIQ_PIS=linha[13]
		            ALIQ_PIS=float(ALIQ_PIS.replace(",", "."))
		            VL_PIS=linha[14]
		            VL_PIS=float(VL_PIS.replace(",", "."))
		            CST_COFINS=linha[15]
		            VL_BC_COFINS=linha[16]
		            VL_BC_COFINS=float(VL_BC_COFINS.replace(",", "."))
		            ALIQ_COFINS=linha[17]
		            ALIQ_COFINS=float(ALIQ_COFINS.replace(",", "."))
		            VL_COFINS=linha[18]
		            VL_COFINS=float(VL_COFINS.replace(",", "."))
		            COD_CTA=linha[16]
		            COD_CCUS=linha[20]
		            DESC_BEM_IMOB=linha[21]
		            
		            #REGRAS
		            # RECALCULO PIS
		            RECALC_PIS_F130=round((VL_BC_PIS * ALIQ_PIS / 100),2)
		            DIF_RECALC_PIS_F130=round(RECALC_PIS_F130-VL_PIS,2)
		            if DIF_RECALC_PIS_F130>0.1 or DIF_RECALC_PIS_F130<-0.1:
		                INCONSISTENCIA='Diferença Cálculo Pis'
		                OBSERV='Valor Sped = ' + str(VL_PIS)+' / ' + 'Valor Recalculado = ' + str(RECALC_PIS_F130)+' / ' + 'Dif = ' + str(DIF_RECALC_PIS_F130)
		                CAMPOS_REL_BLOCO_F130()
		                INCONSISTENCIA=''
		                OBSERV=''
		            # RECALCULO COFINS
		            RECALC_COFINS_F130=round((VL_BC_COFINS * ALIQ_COFINS / 100),2)
		            DIF_RECALC_COFINS_F130=round(RECALC_COFINS_F130-VL_COFINS,2)
		            if DIF_RECALC_COFINS_F130>0.1 or DIF_RECALC_COFINS_F130<-0.1:
		                INCONSISTENCIA='Diferença Cálculo COfins'
		                OBSERV='Valor Sped = ' + str(VL_COFINS)+' / ' + 'Valor Recalculado = ' + str(RECALC_COFINS_F130) +' / ' + 'Dif = ' + str(DIF_RECALC_COFINS_F130)
		                CAMPOS_REL_BLOCO_F130()
		                INCONSISTENCIA=''
		                OBSERV=''
		            # ORIGEM BEM X ALIQUOTA
		            if IND_ORIG_CRED=='0' and ALIQ_PIS!=1.65 or IND_ORIG_CRED=='0' and ALIQ_COFINS!=7.6:
		                INCONSISTENCIA='Bem Nacional com Aliquota Divergente'
		                OBSERV='Verifique se a aliquota está correta para bem nacional'
		                CAMPOS_REL_BLOCO_F130()
		                INCONSISTENCIA=''
		                OBSERV=''
		            if IND_ORIG_CRED=='1' and ALIQ_PIS==1.65 or IND_ORIG_CRED=='1' and ALIQ_COFINS==7.6:
		                INCONSISTENCIA='Bem Importado com Aliquota Divergente'
		                OBSERV='Verifique se a aliquota está correta para bem importado'
		                CAMPOS_REL_BLOCO_F130()
		                INCONSISTENCIA=''
		                OBSERV='' 
		            # CST X ALIQUOTA
		            if CST_PIS==56 and ALIQ_PIS==0 or CST_PIS==56 and ALIQ_COFINS==0 :
		                INCONSISTENCIA='Aliquota incorreta para CST gerador de crédito'
		                OBSERV=''
		                CAMPOS_REL_BLOCO_F130()
		                INCONSISTENCIA=''
		                OBSERV=''
		    
		            
		  # MONTA RELATÓRIO FINAL - BLOCO_D - Cabeçalhos
		    RELATORIO_BLOCO_F130=pd.DataFrame(RELATORIO_BLOCO_F130,columns=['INCONSISTÊNCIA','OBSERVACAO','REG','CNPJ',\
		                                                                    'IND_ORIG_CRED','DESC_BEM_IMOB',\
		                                                                    'VL_BC_COFINS','ALIQ_COFINS','VL_COFINS','VL_BC_PIS',\
		                                                                    'ALIQ_PIS','VL_PIS'])
		    # Remove linhas duplicadas
		    RELATORIO_BLOCO_F130=RELATORIO_BLOCO_F130.drop_duplicates()
		    #RELATORIO_BLOCO_F130.to_excel('c:\\temp\\' + NOME_EXCEL + '.xlsx',encoding='utf-8', index=True )
		    RELATORIO_BLOCO_F130.to_excel(PLANILHA, sheet_name='BLOCO_F130',index=False)
		   # PLANILHA.save()
		    # Descreve na tela a etapa do processo atual....
		    
		    #---------------------------------------------------------------------------------------------------------
		    # REGISTRO 02000 - Descrições iguais para itens diferentes
		    
		    # Define nome da planilha final e zera a lista
		    NOME_EXCEL='INCONS_EFD_CONTRIBUIÇÕES_BLOCO_0200'
		    RELATORIO_BLOCO_0200=[]
		    # cria função das informações que vão aparecer no relatório
		    def CAMPOS_REL_BLOCO_0200():
		        RELATORIO_BLOCO_0200.append((INCONSISTENCIA,OBSERV,COD_ITEM,DESCR_ITEM,TIPO_ITEM,DESCR_TIPO_ITEM,COD_GEN))

		  
		    for indice,linha in LISTA_REG_0200.iterrows():
		        COD_ITEM=linha[0]
		        DESCR_ITEM=linha[1]
		        TIPO_ITEM=linha[2]
		        COD_GEN=linha[6]
		        #st.write(COD_GEN)
		        contador=0
		        if TIPO_ITEM=='00':
		        	DESCR_TIPO_ITEM='Mercadoria para Revenda'
		        if TIPO_ITEM=='01':
		        	DESCR_TIPO_ITEM='Matéria-Prima'
		        if TIPO_ITEM=='02':
		        	DESCR_TIPO_ITEM='Embalagem'
		        if TIPO_ITEM=='03':
		        	DESCR_TIPO_ITEM='Produto em Processo'
		        if TIPO_ITEM=='04':
		        	DESCR_TIPO_ITEM='Produto Acabado'
		        if TIPO_ITEM=='05':
		        	DESCR_TIPO_ITEM='Subproduto'
		        if TIPO_ITEM=='06':
		        	DESCR_TIPO_ITEM='Produto Intermediário'
		        if TIPO_ITEM=='07':
		        	DESCR_TIPO_ITEM='Material de Uso e Consumo'
		        if TIPO_ITEM=='08':
		        	DESCR_TIPO_ITEM='Ativo Imobilizado'
		        if TIPO_ITEM=='09':
		        	DESCR_TIPO_ITEM='Serviços'
		        if TIPO_ITEM=='10':
		        	DESCR_TIPO_ITEM='Outros insumos'
		        if TIPO_ITEM=='99':
		        	DESCR_TIPO_ITEM='Outras'
		        for indice,linha in LISTA_REG_0200.iterrows():
		            if  COD_ITEM!=linha[0] and DESCR_ITEM==linha[1] or COD_ITEM==linha[0] and DESCR_ITEM!=linha[1]:
		                contador=1
		            if COD_ITEM!=linha[0] and DESCR_ITEM==linha[1] or COD_ITEM==linha[0] and DESCR_ITEM!=linha[1]:
		                INCONSISTENCIA='Descrições iguais para itens diferentes ou Itens iguais para descrição Diferente'
		                OBSERV
		                CAMPOS_REL_BLOCO_0200()
		                INCONSISTENCIA=''
		                OBSERV=''
		            if COD_GEN=='00' and TIPO_ITEM!='09':
		            	INCONSISTENCIA='Tipo do Item Incompatível com o Código do gênero informado (00 - Serviço).'
		            	OBSERV='TIPO_ITEM - ' + TIPO_ITEM + '-' + DESCR_TIPO_ITEM 
		            	CAMPOS_REL_BLOCO_0200()
		            	INCONSISTENCIA=''
		            	OBSERV=''
		                
		    # MONTA RELATÓRIO FINAL - BLOCO_D - Cabeçalhos
		    RELATORIO_BLOCO_0200=pd.DataFrame(RELATORIO_BLOCO_0200,columns=['INCONSISTÊNCIA','OBSERVACAO','COD_ITEM',\
		    	'DESCR_ITEM','TIPO_ITEM','DESCR_TIPO_ITEM','COD_GEN'])
		    # Remove linhas duplicadas
		    RELATORIO_BLOCO_0200=RELATORIO_BLOCO_0200.drop_duplicates()
		    #RELATORIO_BLOCO_0200.to_excel('c:\\temp\\' + NOME_EXCEL + '.xlsx',encoding='utf-8', index=True )
		    RELATORIO_BLOCO_0200.to_excel(PLANILHA, sheet_name='BLOCO_0200',index=False)
		    #PLANILHA.save()
		            
		   
		    # Cruzamento com Mapa de Vendas
		    # Prepara tabela do preço unitario das saidas do sped
		

		    # salva relatórios
		    
		    # Grava uma cópia do resumo na pasta da extração do sped
		    EXTRACAO = pd.ExcelWriter('G:\\##Obrig Legais#\\RFB -Sped Pis-Cof\\Extração_Sped\\' \
		    	+'EFD_CONTR_'+str(CNPJ1)+'-'+ str(DT_INI) + ' a ' + str(DT_FIN) +".xlsx", engine='xlsxwriter')

		    MONTAGEM_RESUMO_C.to_excel(EXTRACAO, sheet_name='RESUMO_NF',index=False)

		    RELATORIO_BLOCO_D.to_excel(EXTRACAO, sheet_name='BLOCO_D',index=False)
		    EXTRACAO.save()
		    EXTRACAO.close()



		    # Grava resumo geral na planilha de auditoria
		    MONTAGEM_RESUMO_C.to_excel(PLANILHA, sheet_name='RESUMO_NF',index=False)
		    #st.dataframe(MONTAGEM_RESUMO_C)
		    #Botão para baixar planilha
		    #st.button("baixar planilha",PLANILHA)


		    # Montagem Resumo para Processo
		    PisCofins_AliqZero = pd.ExcelWriter('c://temp//'+'PisCofins_AliqZero_'+str(CNPJ1)+'-'+ str(DT_INI) + ' a ' + str(DT_FIN) +'.xlsx', engine='xlsxwriter')
		    REL_ZERO=MONTAGEM_RESUMO_C.loc[((MONTAGEM_RESUMO_C['IND_OPER']) == 'Saída')]
		    REL_ZERO=REL_ZERO.loc[((REL_ZERO['CST_PIS']) == '06')]
		    #REL_NOTA=REL_NOTA.loc[((REL_NOTA['VL_ICMS']) !=0)]
		    REL_NOTA=REL_ZERO[['CNPJ','DT_DOC','DT_E_S','NUM_DOC','SER','VL_DOC']]
		    REL_ITEM=REL_ZERO[['CNPJ','DT_DOC','DT_E_S','NUM_DOC','SER','COD_ITEM','DESCR_ITEM','VL_ITEM','VL_ICMS']]
		    REL_NOTA.drop_duplicates(inplace=True)
		    #REL_NOTA['VL_DOC']=REL_NOTA['VL_DOC'].str.replace(',','.')
		    #REL_ITEM['VL_ITEM']=REL_ITEM['VL_ITEM'].str.replace(',','.')
		    #REL_ITEM['VL_ICMS']=REL_ITEM['VL_ICMS'].str.replace(',','.')

		    #REL_NOTA['VL_DOC']=REL_NOTA['VL_DOC'].astype(float)
		    #REL_ITEM['VL_ITEM']=REL_ITEM['VL_ITEM'].astype(float)
		    #REL_ITEM['VL_ICMS']=REL_ITEM['VL_ICMS'].astype(float)
		    REL_NOTA.to_excel(PisCofins_AliqZero, sheet_name='ALIQUOTA_ZERO_NOTA',index=False)
		    REL_ITEM.to_excel(PisCofins_AliqZero, sheet_name='ALIQUOTA_ZERO_ITEM',index=False)


		    PisCofins_AliqZero.save()
		    PisCofins_AliqZero.close()

		    PLANILHA.save()
		    PLANILHA.close()

		    #time.sleep(0.10)
#---------------------------------------------------------------------------------------------------------------------------------------------------------
		    # Formata Planilha
		    #wb = xl.load_workbook(PLANILHA, read_only=False, keep_vba=False)
		    #ws = wb.sheetnames
		    
		    # Formata como tabela
		    #AbasFormatacao=['BLOCO_C','BLOCO_D','BLOCO_C500','BLOCO_F120','BLOCO_F130','BLOCO_0200','RESUMO_NF']
		    #for i in AbasFormatacao:
		    	#ws = wb[i]

		    	#if ws['a2']!=None:
			    	 # pega endereço da coluna
			    	#rows = get_column_letter(ws.max_column) + str(ws.max_row)
			    	#cols = ws.max_column
			    	#intervalo='a1' + ':' + rows

			    	#tab = Table(displayName=i, ref=intervalo)
			    	#style = TableStyleInfo(name="TableStyleMedium16", showFirstColumn=False,
		                       #showLastColumn=False, showRowStripes=True, showColumnStripes=False)
			    	#tab.tableStyleInfo = style
			    	#ws.add_table(tab)

		    #wb.save(PLANILHA)
		    #wb.close()
#---------------------------------------------------------------------------------------------------------------------------------------------------------

		   
		    st.success(':white_check_mark: Fim de Auditoria. Arquivo Salvo em G:\#Sistemas -ERP etc\# AUXILIUS ')
		    
		 

	if st.button('Auditar'):
		ExecutaAuditoria()